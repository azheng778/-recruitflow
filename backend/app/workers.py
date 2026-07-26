from __future__ import annotations

import asyncio
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import or_, select

from .config import settings
from .database import SessionLocal
from .models import DocumentSyncJob, Notification, utcnow


EXCEL_HEADERS = [
    "entity_id",
    "candidate_id",
    "candidate_name",
    "job_id",
    "job_name",
    "status",
    "owner_id",
    "next_action",
    "next_action_at",
    "updated_at",
]


def _write_excel(job: DocumentSyncJob) -> str:
    settings.export_dir.mkdir(parents=True, exist_ok=True)
    target = settings.export_dir / "recruitment_data.xlsx"
    if target.exists():
        workbook = load_workbook(target)
        sheet = workbook["Recruitment"] if "Recruitment" in workbook.sheetnames else workbook.create_sheet("Recruitment")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Recruitment"
        sheet.append(EXCEL_HEADERS)
        sheet.freeze_panes = "A2"
    entity_id = job.entity_id
    row_number = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == entity_id:
            row_number = row
            break
    if row_number is None:
        row_number = sheet.max_row + 1
    values = {"entity_id": entity_id, **job.payload}
    for column, header in enumerate(EXCEL_HEADERS, start=1):
        value = values.get(header)
        if isinstance(value, (list, dict)):
            value = str(value)
        sheet.cell(row=row_number, column=column, value=value)
    workbook.save(target)
    return f"Recruitment!{row_number}"


def run_worker_once() -> dict[str, int]:
    processed_sync = 0
    processed_notifications = 0
    now = utcnow()
    with SessionLocal() as db:
        jobs = db.scalars(
            select(DocumentSyncJob)
            .where(
                DocumentSyncJob.status == "pending",
                or_(DocumentSyncJob.next_retry_at.is_(None), DocumentSyncJob.next_retry_at <= now),
            )
            .order_by(DocumentSyncJob.created_at)
            .limit(20)
        ).all()
        for job in jobs:
            job.status = "running"
            job.last_attempt_at = now
            db.commit()
            try:
                if job.sink_type == "local_excel":
                    job.external_record_id = _write_excel(job)
                elif job.sink_type == "tencent_docs" and not settings.tencent_docs_enabled:
                    raise RuntimeError("Tencent Docs integration is disabled")
                else:
                    raise RuntimeError(f"Unsupported sink: {job.sink_type}")
                job.status = "succeeded"
                job.completed_at = utcnow()
                job.error_code = None
                job.error_message = None
                processed_sync += 1
            except Exception as exc:
                job.retry_count += 1
                job.error_code = type(exc).__name__
                job.error_message = str(exc)[:500]
                if job.retry_count >= 3:
                    job.status = "failed"
                    job.next_retry_at = None
                else:
                    job.status = "pending"
                    delay = [1, 5, 15][job.retry_count - 1]
                    job.next_retry_at = utcnow() + timedelta(minutes=delay)
            db.commit()

        notifications = db.scalars(
            select(Notification)
            .where(Notification.status == "pending", Notification.scheduled_at <= now)
            .order_by(Notification.scheduled_at)
            .limit(20)
        ).all()
        for item in notifications:
            item.status = "sent"
            item.sent_at = utcnow()
            processed_notifications += 1
        db.commit()
    return {"sync": processed_sync, "notifications": processed_notifications}


async def worker_loop(stop_event: asyncio.Event, interval_seconds: int = 10) -> None:
    while not stop_event.is_set():
        await asyncio.to_thread(run_worker_once)
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=interval_seconds)
        except TimeoutError:
            continue

